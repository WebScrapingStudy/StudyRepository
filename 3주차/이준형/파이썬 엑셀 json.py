#!/usr/bin/env python
# coding: utf-8

# In[4]:


#파일을 읽어들이는 내장함수인 load_workbook을 불러옵니다.
from openpyxl import load_workbook
#파일이름이 angel.xlsx인 파일을 불러옵니다.
angelEx=load_workbook(filename='angel.xlsx')

 

#불러온 엑셀 파일 중 데이터를 찾을 sheet의 이름을 입력합니다.
sheet1 = angelEx['Sheet1']
#활성화되어 있는 시트를 찾습니다.
sheet2 = angelEx.active
#Sheet1의 D4의 값을 출력합니다.
print(sheet1['D4'].value)
print(sheet2['D4'].value)

#특정 범위의 데이터(번호와 성인명)만 불러오겠습니다.
angels = []
#루프문을 이용해 sheet의 여러 행에 있는 데이터를 불러옵니다.
for i in sheet1.rows:
    번호 = i[0].value
    성인명 = i[2].value
    angel = (번호, 성인명)
    angels.append(angel)

print(angels)


# In[7]:


from openpyxl import Workbook
wb = Workbook()


#파일 이름을 정하고, 데이터를 넣을 시트를 활성화합니다.
sheet1 = wb.active
file_name = 'sample.xlsx'

#시트의 이름을 정합니다.
sheet1.title = 'sampleSheet'

abc = ['A','B','C','D','E','F','G','H','I','J','K','L']

 


#cell 함수를 이용해 넣을 데이터의 행렬 위치를 지정해줍니다.
for row_index in range(1, 11):
    sheet1.cell(row=row_index, column=1).value = row_index
    sheet1.cell(row=row_index, column=2).value = abc[(row_index-1)]

wb.save(filename=file_name)


# In[6]:


import openpyxl
import codecs

# 엑셀파일 열기
filename = "word.xlsx"
book = openpyxl.load_workbook(filename)

# 엑셀 파일의 첫번째 시트 추출하기
sheet = book.worksheets[0]

# json 형식에 맞게 문자 변경
def replaceToJson(text):
	if text == None:
		return ""
	else:
		text = text.replace("\n", "\\n")
		text = text.replace('"', '\\"')
		return text

# 저장할 json 파일 오픈
jsonFile = codecs.open('word.json', 'w',  'utf-8')
jsonFile.write('{"eng": [' + '\n')

# 시트의 각 행을 순서대로 추출하기
i = 1
max_row = sheet.max_row

for row in sheet.rows:
	word	= replaceToJson(row[0].value)
	means	= replaceToJson(row[1].value)
	example	= replaceToJson(row[2].value)

	if i != max_row:
		jsonword = '\t{"word": "%s", "korean": "%s", "example": "%s"},' % (word, means, example)
	else:
		jsonword = '\t{"word": "%s", "korean": "%s", "example": "%s"}' % (word, means, example)

	# json 파일에 쓰기
	jsonFile.write(jsonword + '\n')
	i = i + 1

jsonFile.write(']}')
jsonFile.close()


# In[ ]:





# In[ ]:




